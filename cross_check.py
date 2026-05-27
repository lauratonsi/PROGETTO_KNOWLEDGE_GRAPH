import openpyxl, csv, unicodedata

def normalize(s):
    if s is None: return ''
    s = str(s).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    for prefix in ['PASSAGGIO ', 'PIAZZALE ', 'VIA ', 'VIALE ', 'PIAZZA ', 'PIAZZETTA ']:
        if s.startswith(prefix):
            s = s[len(prefix):]
            break
    s = s.replace("'", '').replace('`', '').replace('(STRADA)', '').strip()
    return s

wb = openpyxl.load_workbook('F_M DATE.xlsx')

# DONNE sheet
ws_d = wb['DONNE']
xlsx_donne = {}
for row in ws_d.iter_rows(min_row=2, values_only=True):
    if row[0] and str(row[0]) != 'Nome Personaggio':
        xlsx_donne[normalize(str(row[0]))] = (row[0], row[2], row[3], row[4])

# INCERTI sheet
ws_i = wb['INCERTI']
incerti_names = set()
for row in ws_i.iter_rows(min_row=2, values_only=True):
    if row[0]:
        incerti_names.add(normalize(str(row[0])))

print('INCERTI identico a DONNE?', incerti_names == set(xlsx_donne.keys()))
print()

# UOMINI sheet
ws_u = wb['UOMINI']
xlsx_uomini = {}
for row in ws_u.iter_rows(min_row=1, values_only=True):
    if row[0]:
        xlsx_uomini[normalize(str(row[0]))] = (row[0], row[2], row[3], row[4])

print(f'XLSX: {len(xlsx_uomini)} uomini, {len(xlsx_donne)} donne')
print()

# CSV
csv_female = []
csv_male = []
with open('bologna_KG_ready.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        g = row['GENERE']
        nome = row['NOME_PULITO'].strip()
        if g == 'Female':
            csv_female.append(nome)
        elif g == 'Male':
            csv_male.append(nome)

print(f'CSV: {len(csv_male)} Male, {len(csv_female)} Female')
print()

print('=== DONNE: CSV vs XLSX ===')
matched = []
partial = []
missing = []
for nome in sorted(csv_female):
    norm = normalize(nome)
    if norm in xlsx_donne:
        matched.append((nome, xlsx_donne[norm][0]))
    else:
        found = None
        for k, v in xlsx_donne.items():
            if norm in k or k in norm:
                found = v[0]
                break
        if found:
            partial.append((nome, found))
        else:
            missing.append(nome)

print(f'Match esatto: {len(matched)}')
for a, b in matched:
    print(f'  OK  {a}  ->  {b}')
print()
print(f'Match parziale: {len(partial)}')
for a, b in partial:
    print(f'  ~   {a}  ->  {b}')
print()
print(f'MANCANTI nel XLSX ({len(missing)}):')
for nome in missing:
    print(f'  MANCA  {nome}')
