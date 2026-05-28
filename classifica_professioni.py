"""
Legge F_M DATE.xlsx (UOMINI + DONNE) e proposte_intitolazioni_future.csv,
assegna una macro-categoria a ogni persona, salva classificazione_professioni.csv
e stampa le statistiche comparative per genere.
"""
import sys, csv, re
from pathlib import Path
import openpyxl
from collections import Counter, defaultdict

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
BASE = Path(__file__).parent

# ── Tassonomia ────────────────────────────────────────────────────────────────
# Ogni categoria: lista di sottostringhe da cercare (lowercase) nella professione.
# L'ordine delle CATEGORIE determina la priorità quando più categorie matchano.
CATEGORIE = [
    ("Resistenza e antifascismo", [
        "partigian", "staffetta", "antifascist", "resistenza", "combattente della resist",
        "caduta per la libera", "perseguit", "deportat",
    ]),
    ("Sport", [
        "atleta", "ciclista", "schermit", "olimpion", "velocista", "nuotatr",
        "calciat", "sportiv", "pugil", "ginnast", "tennist", "maratonet",
        "saltator", "lanciat", "giavellott", "pilota", "motociclista",
        "automobilista", "alpinist", "scalator", "arrampic", "sciator",
        "remator", "vogator", "pattinator", "boxeur", "lottator",
    ]),
    ("Religione", [
        "sacerdote", "presbitero", "suora", "monaco", "monaca", "religios",
        "mistic", "missionari", "abate", "vescov", "cardinale", "pap",
        "fondatrice del movimento dei focolari", "fondatric.*ordine",
        "santa,", "santa ", "sante",
    ]),
    ("Istruzione ed educazione", [
        "maestra", "insegnant", "pedagogist", "educatr", "educatore",
        "pioniera pedagogia", "direttrice di coro", "tiflodidatta",
        "preside", "rettore", "rettrice",
    ]),
    ("Sindacalismo e attivismo civile", [
        "sindacalist", "attivista", "emancipazionist", "dirigente sindac",
        "attivista trans", "attivista lgbtq", "attivista diritti",
        "militante polit", "funzionaria polit", "difensore",
        "paladina", "filantropa", "benefattr", "filantropo",
        "giusto tra le nazioni", "giusto fra le nazioni",
    ]),
    ("Politica e diritto", [
        "politico", "politica", "avvocato", "avvocata", "giurista",
        "diplomat", "sindaco", "parlament", "senatore", "senatrice",
        "deputat", "ministr", "consiglier", "governator", "prefetto",
        "pubblico amministr", "magistrat", "procurator", "giudice",
        "economi", "economista", "statistico", "statistica",
        "pres. repubblica", "presidente della republic",
    ]),
    ("Scienze e medicina", [
        "medico", "medica", "scienziata", "scienziato", "fisico", "fisica",
        "matematico", "matematica", "chirurgo", "chirurga", "chimica", "chimico",
        "anatomist", "biolog", "astronom", "botan", "fisiolog",
        "naturalista", "ricercat", "batteriolog", "farmac", "geolog",
        "zoolog", "limnolog", "radiolog", "pediatra", "ginecologo", "ginecologa",
        "ostetrica", "ostetrico", "ingegner", "psicoanalista", "psicolog",
        "psichiatr", "neuroscient", "oftalm", "otorinolaring",
        "cardiolog", "patologo", "patologa", "oculi", "pneumolog",
        "oncologo", "dermatologo", "cosmonauta", "astronauta",
        "clinico", "clinica", "tossicologo", "epidemiolog",
        "inventore", "inventrice",
    ]),
    ("Filosofia, storia e accademia", [
        "filosofo", "filosofa", "storico", "storica", "professore univers",
        "professoressa", "accademico", "accademica", "archeolog",
        "linguista", "umanista", "filolog", "medievist", "letterato", "letterata",
        "erudita", "erudito", "critico letter", "critica letter",
        "storico dell'arte", "storica dell'arte", "critica d'arte", "critico d'arte",
        "paleograf", "storico med", "storica med", "glottoteta", "semiologo",
        "antropolog", "sociologo", "sociologa", "latinista", "dantista",
        "dialettologo", "grecista", "orientalist", "indolog", "slavist",
        "romanist", "germanist", "byzantinist", "cancelliere",
    ]),
    ("Arte visiva e architettura", [
        "pittore", "pittrice", "scultore", "scultrice", "incisore",
        "fotografo", "fotografa", "orafo", "orafessa", "ceramist",
        "architetto", "architetta", "urbanist", "designer",
        "ceroplast", "miniaturist", "disegnat", "fumettist", "vignettist",
        "illustrat", "xilograf", "acquerellista", "mosaicist",
    ]),
    ("Musica, teatro e cinema", [
        "compositore", "compositrice", "musicist", "cantante", "soprano",
        "tenor", "barit", "mezzosopran", "attore", "attrice", "regist",
        "drammaturgo", "drammaturga", "commediografo", "commediografa",
        "direttore d'orchestra", "direttrice d'orchestra", "violinist",
        "pianist", "organist", "cantante liric", "documentarist",
        "coreograf", "ballerina", "ballerino", "cantautore", "cantautrice",
        "violoncellist", "liutaio", "liutaia", "oboist", "flautist",
        "trombonist", "cornist", "chitarrist", "arpist", "cembalista",
        "impresario teatr", "scenograf",
    ]),
    ("Letteratura e giornalismo", [
        "scrittore", "scrittrice", "poeta", "poetessa", "giornalist",
        "traduttr", "traduttore", "editore", "editrice", "bibliotecari",
        "saggist", "romanzier", "narratore", "narratrice",
    ]),
    ("Patrioti, militari ed esploratori", [
        "patriota", "militare", "generale", "esploratore", "esploratrice",
        "aviatore", "aviatrice", "aviaz", "ammiraglio", "capitano", "colonnello",
        "soldato", "guerriero", "condottiero", "console roman",
        "tribuno", "imperatore",
    ]),
]

ALTRO = "Altro / istituzionale"


def classifica(professione: str) -> str:
    if not professione or professione.strip() in ('', 'N.D.', 'N.d.', 'n.d.'):
        return ALTRO
    p = professione.lower()
    for nome_cat, keywords in CATEGORIE:
        for kw in keywords:
            if re.search(kw, p):
                return nome_cat
    return ALTRO


# ── Lettura dati ──────────────────────────────────────────────────────────────
persone = []   # (nome, genere, tipo, professione_orig, macro_cat)

wb = openpyxl.load_workbook(BASE / 'F_M DATE.xlsx')

HEADER_VALS = {'nome personaggio', 'nome', 'name', 'professione / ruolo', 'professione'}

for row in wb['UOMINI'].iter_rows(min_row=2, values_only=True):
    nome, genere, prof = row[0], row[1], row[2]
    if not nome or str(nome).strip().lower() in HEADER_VALS:
        continue
    prof = str(prof) if prof else ''
    persone.append((str(nome), 'M', 'storica', prof, classifica(prof)))

for row in wb['DONNE'].iter_rows(min_row=2, values_only=True):
    nome, genere, prof = row[0], row[1], row[2]
    if not nome or str(nome).strip().lower() in HEADER_VALS:
        continue
    prof = str(prof) if prof else ''
    persone.append((str(nome), 'F', 'storica', prof, classifica(prof)))

with open(BASE / 'proposte_intitolazioni_future.csv', encoding='utf-8') as f:
    for r in csv.DictReader(f):
        prof = r.get('professione', '')
        persone.append((r['nome_completo'], 'F', 'proposta', prof, classifica(prof)))

# ── Salva CSV ─────────────────────────────────────────────────────────────────
out = BASE / 'classificazione_professioni.csv'
with open(out, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['nome', 'genere', 'tipo', 'professione_originale', 'macro_categoria'])
    w.writerows(persone)
print(f'Salvato: {out.name}  ({len(persone)} righe)')

# ── Statistiche ───────────────────────────────────────────────────────────────
uomini  = [p for p in persone if p[1] == 'M']
donne   = [p for p in persone if p[1] == 'F' and p[2] == 'storica']
proposte= [p for p in persone if p[2] == 'proposta']

def pct_dist(lista):
    cnt = Counter(p[4] for p in lista)
    tot = len(lista)
    return {k: (v, round(v/tot*100, 1)) for k,v in sorted(cnt.items(), key=lambda x: -x[1])}

print(f'\n{"="*65}')
print(f'TOTALE: {len(persone)}  |  Uomini: {len(uomini)}  |  Donne storiche: {len(donne)}  |  Proposte: {len(proposte)}')
print(f'{"="*65}')

cats = [c[0] for c in CATEGORIE] + [ALTRO]
header = f'{"Categoria":<38} {"TOT":>5}  {"M %":>6}  {"F %":>6}  {"PROP %":>7}'
print('\n' + header)
print('-'*65)

all_cnt  = Counter(p[4] for p in persone)
m_cnt    = Counter(p[4] for p in uomini)
f_cnt    = Counter(p[4] for p in donne)
pr_cnt   = Counter(p[4] for p in proposte)

for cat in cats:
    tot = all_cnt.get(cat, 0)
    if tot == 0:
        continue
    mp  = f'{m_cnt.get(cat,0)/len(uomini)*100:.1f}%' if uomini else '-'
    fp  = f'{f_cnt.get(cat,0)/len(donne)*100:.1f}%' if donne else '-'
    prp = f'{pr_cnt.get(cat,0)/len(proposte)*100:.1f}%' if proposte else '-'
    print(f'{cat:<38} {tot:>5}  {mp:>6}  {fp:>6}  {prp:>7}')

print('-'*65)
print(f'{"TOTALE":<38} {len(persone):>5}  {"100%":>6}  {"100%":>6}  {"100%":>7}')

# Casi non classificati
nd = [p for p in persone if p[4] == ALTRO]
if nd:
    print(f'\nNon classificati ({len(nd)}):')
    for p in nd[:20]:
        print(f'  [{p[1]}] {p[0][:40]:42} | {p[3][:50]}')
    if len(nd) > 20:
        print(f'  ... e altri {len(nd)-20}')
