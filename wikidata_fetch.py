"""
Fetch Wikidata data (occupation, birth, death) for all persons
missing from F_M DATE.xlsx and append them to the correct sheet.

Run:  python wikidata_fetch.py
Results cached in wikidata_cache.json so the script is resumable.
Persons not found on Wikidata are written to wikidata_notfound.txt.
"""

import csv, json, time, unicodedata, re, sys
import requests
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASE = Path(__file__).parent
CSV_IN      = BASE / 'bologna_KG_ready.csv'
XLSX_IN     = BASE / 'F_M DATE.xlsx'
CACHE_FILE  = BASE / 'wikidata_cache.json'
NOTFOUND    = BASE / 'wikidata_notfound.txt'

# Original row counts (before our additions)
UOMINI_ORIGINAL_ROWS = 654
DONNE_ORIGINAL_DATA_ROWS = 40   # header is row 1, data rows 2-41

MONTHS_IT = ['', 'gennaio', 'febbraio', 'marzo', 'aprile', 'maggio', 'giugno',
             'luglio', 'agosto', 'settembre', 'ottobre', 'novembre', 'dicembre']

STREET_PREFIXES = [
    'PASSAGGIO ', 'PIAZZALE ', 'PIAZZETTA ', 'PIAZZA ', 'VIALE ', 'VIA ',
    'ROTONDA ', 'GALLERIA ', 'PONTE ', 'LARGO ', "LOCALITA' ",
    'MURA ', 'SOTTOPASSO ', 'VICOLO ', 'SALITA ',
]

COLLECTIVES = ['FRATELLI ', 'RAGAZZI DEL ']

# ---------------------------------------------------------------------------
# Name helpers
# ---------------------------------------------------------------------------

def strip_street_prefix(s):
    for p in STREET_PREFIXES:
        if s.startswith(p):
            return s[len(p):]
    return s

def normalize_key(s):
    s = str(s).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    s = strip_street_prefix(s)
    s = s.replace("'", '').replace('`', '').replace('.', '').replace('-', ' ')
    return re.sub(r'\s+', ' ', s).strip()

PARTICLES = {
    'DI', 'DA', 'DE', 'DEL', 'DELLA', 'DEGLI', 'DELLE', 'DELLO',
    'DAL', 'DALLA', 'DALL', 'D', 'E', 'ED', 'IN', 'IL', 'LA', 'LO',
    'I', 'LE', 'LI', 'UN', 'UNA', 'AND', 'OF', 'THE',
}

def to_search_name(csv_name):
    s = strip_street_prefix(csv_name.strip().upper())
    words = s.split()
    result = []
    for i, w in enumerate(words):
        if "'" in w:
            parts = w.split("'", 1)
            result.append(parts[0].capitalize() + "'" + parts[1].capitalize())
        elif i > 0 and w in PARTICLES:
            result.append(w.lower())
        else:
            result.append(w.capitalize())
    return ' '.join(result)

def is_collective(csv_name):
    return any(csv_name.startswith(c) for c in COLLECTIVES)

# ---------------------------------------------------------------------------
# Wikidata API (with retry)
# ---------------------------------------------------------------------------

SESSION = requests.Session()
SESSION.headers.update({
    'User-Agent': 'BolognaKGBot/1.0 university project; https://github.com/lauratonsi/PROGETTO_KNOWLEDGE_GRAPH'
})

def _get_json(url, params, retries=3):
    for attempt in range(retries):
        try:
            r = SESSION.get(url, params=params, timeout=20)
            if r.status_code == 429:
                wait = 10 * (attempt + 1)
                print(f'    Rate limited, waiting {wait}s ...')
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json()
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(3 * (attempt + 1))
            else:
                raise e
    return {}

def search_entity(name):
    for lang in ('it', 'en'):
        try:
            data = _get_json('https://www.wikidata.org/w/api.php', {
                'action': 'wbsearchentities',
                'search': name,
                'language': lang,
                'type': 'item',
                'limit': 5,
                'format': 'json',
            })
            hits = data.get('search', [])
            if hits:
                return hits
        except Exception as e:
            print(f'    [search error] {name} ({lang}): {e}')
    return []

def fetch_entities_batch(qids, retries=4):
    for attempt in range(retries):
        try:
            data = _get_json('https://www.wikidata.org/w/api.php', {
                'action': 'wbgetentities',
                'ids': '|'.join(qids),
                'props': 'claims|labels',
                'languages': 'it|en',
                'format': 'json',
            })
            return data.get('entities', {})
        except Exception as e:
            wait = 5 * (attempt + 1)
            print(f'    [batch fetch error attempt {attempt+1}]: {e} -> retry in {wait}s')
            time.sleep(wait)
    return {}

def is_human(claims):
    """True=human, False=not human, None=unknown (empty entity)."""
    if not claims:
        return None
    p31 = claims.get('P31', [])
    if not p31:
        return None  # can't tell without P31
    for c in p31:
        v = c.get('mainsnak', {}).get('datavalue', {}).get('value', {})
        if isinstance(v, dict) and v.get('id') == 'Q5':
            return True
    return False

def get_entity_label(ent):
    for lang in ('it', 'en'):
        lbl = ent.get('labels', {}).get(lang, {}).get('value')
        if lbl:
            return lbl
    return ''

def format_time(tv):
    try:
        t    = tv['time']
        prec = tv['precision']
        bce  = t.startswith('-')
        bare = t.lstrip('+-').split('T')[0]
        parts = bare.split('-')
        year = parts[0].lstrip('0') or '0'
        if bce:
            year += ' a.C.'
        if prec >= 11 and len(parts) == 3:
            m, d = int(parts[1]), int(parts[2])
            if m > 0 and d > 0:
                return f'{d} {MONTHS_IT[m]} {year}'
        if prec >= 10 and len(parts) >= 2:
            m = int(parts[1])
            if m > 0:
                return f'{MONTHS_IT[m]} {year}'
        return year
    except Exception:
        return ''

def extract_person_data(entity, place_labels):
    claims = entity.get('claims', {})

    occs = []
    for c in claims.get('P106', []):
        qid = c.get('mainsnak', {}).get('datavalue', {}).get('value', {}).get('id')
        if qid and qid in place_labels:
            occs.append(place_labels[qid])
    occupation = ', '.join(occs[:2]) if occs else ''

    def date_place(date_prop, place_prop):
        date_str  = ''
        place_str = ''
        dp = claims.get(date_prop, [])
        if dp:
            tv = dp[0].get('mainsnak', {}).get('datavalue', {}).get('value')
            if isinstance(tv, dict):
                date_str = format_time(tv)
        pp = claims.get(place_prop, [])
        if pp:
            qid = pp[0].get('mainsnak', {}).get('datavalue', {}).get('value', {}).get('id')
            if qid and qid in place_labels:
                place_str = place_labels[qid]
        if date_str and place_str:
            return f'{date_str}, {place_str}'
        return date_str or place_str

    birth = date_place('P569', 'P19') or 'N.D.'
    death = date_place('P570', 'P20') or 'N.D.'
    if not occupation:
        occupation = 'N.D.'

    return occupation, birth, death

# ---------------------------------------------------------------------------
# Step 0 -- clean up any rows added by a previous (failed) run
# ---------------------------------------------------------------------------

print('Opening XLSX and cleaning up previous run ...')
wb = openpyxl.load_workbook(XLSX_IN)

ws_m = wb['UOMINI']
ws_f = wb['DONNE']

# Truncate UOMINI to original row count
current_m = ws_m.max_row
if current_m > UOMINI_ORIGINAL_ROWS:
    for r in range(UOMINI_ORIGINAL_ROWS + 1, current_m + 1):
        for c in range(1, 7):
            ws_m.cell(r, c).value = None
    print(f'  Cleared UOMINI rows {UOMINI_ORIGINAL_ROWS+1}-{current_m}')

# Truncate DONNE to original row count (header + 40 data rows = 41 rows total)
donne_last_original = DONNE_ORIGINAL_DATA_ROWS + 1  # row 41
current_f = ws_f.max_row
if current_f > donne_last_original:
    for r in range(donne_last_original + 1, current_f + 1):
        for c in range(1, 6):
            ws_f.cell(r, c).value = None
    print(f'  Cleared DONNE rows {donne_last_original+1}-{current_f}')

# Fix 'Anna Frank' -> 'Anne Frank' in DONNE (in case it needs doing)
for row in ws_f.iter_rows(min_row=2, max_col=1):
    if row[0].value == 'Anna Frank':
        row[0].value = 'Anne Frank'
        print("  Fixed 'Anna Frank' -> 'Anne Frank'")

wb.save(XLSX_IN)
print('  XLSX cleaned and saved.')

# Reload fresh
wb = openpyxl.load_workbook(XLSX_IN)
ws_m = wb['UOMINI']
ws_f = wb['DONNE']

# ---------------------------------------------------------------------------
# Step 1 -- build missing-persons list
# ---------------------------------------------------------------------------

def load_xlsx_names(wb):
    names = set()
    for sheet, start in (('UOMINI', 1), ('DONNE', 2)):
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=start, values_only=True):
            if row[0] and str(row[0]) not in ('Nome Personaggio',):
                names.add(normalize_key(str(row[0])))
    return names

existing = load_xlsx_names(wb)

missing_m, missing_f = [], []
with open(CSV_IN, encoding='utf-8') as f:
    for row in csv.DictReader(f):
        g    = row['GENERE']
        nome = row['NOME_PULITO'].strip()
        if g not in ('Male', 'Female'):
            continue
        if is_collective(nome):
            continue
        key = normalize_key(nome)
        if key not in existing:
            if g == 'Male':
                missing_m.append(nome)
            else:
                missing_f.append(nome)

print(f'Missing -> Male: {len(missing_m)}, Female: {len(missing_f)}')
if not missing_m and not missing_f:
    print('Nothing to add. Done.')
    sys.exit(0)

# ---------------------------------------------------------------------------
# Step 2 -- search Wikidata (with cache)
# ---------------------------------------------------------------------------

cache = {}
if CACHE_FILE.exists():
    cache = json.loads(CACHE_FILE.read_text(encoding='utf-8'))
    print(f'Cache loaded: {len(cache)} entries')

all_missing = [(n, 'M') for n in missing_m] + [(n, 'F') for n in missing_f]
qid_map = {}
unfound = []

total = len(all_missing)
for i, (nome, gender) in enumerate(all_missing):
    key = normalize_key(nome)
    search_name = to_search_name(nome)

    if key in cache:
        entry = cache[key]
        if entry.get('qid'):
            qid_map[nome] = entry['qid']
        else:
            unfound.append((nome, gender))
        continue

    print(f'  [{i+1}/{total}] Searching: {search_name}')
    time.sleep(0.4)
    hits = search_entity(search_name)

    found_qid = hits[0]['id'] if hits else None
    cache[key] = {'qid': found_qid, 'name': search_name}
    CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding='utf-8')

    if found_qid:
        qid_map[nome] = found_qid
    else:
        unfound.append((nome, gender))

print(f'Search done. Found QIDs: {len(qid_map)}, Not found: {len(unfound)}')

# ---------------------------------------------------------------------------
# Step 3 -- fetch entities in small batches with rate-limit awareness
# ---------------------------------------------------------------------------

BATCH_SIZE = 10
BATCH_DELAY = 2.0   # seconds between batches

qids_needed = list(set(qid_map.values()))
entities = {}
total_batches = (len(qids_needed) + BATCH_SIZE - 1) // BATCH_SIZE

for i in range(0, len(qids_needed), BATCH_SIZE):
    batch = qids_needed[i:i+BATCH_SIZE]
    bn = i // BATCH_SIZE + 1
    print(f'  Fetching entity batch {bn}/{total_batches} ({len(batch)} QIDs) ...')
    ents = fetch_entities_batch(batch)
    entities.update(ents)
    time.sleep(BATCH_DELAY)

# Collect place/occupation QIDs
place_qids = set()
for ent in entities.values():
    for prop in ('P106', 'P19', 'P20'):
        for c in ent.get('claims', {}).get(prop, []):
            v = c.get('mainsnak', {}).get('datavalue', {}).get('value')
            if isinstance(v, dict) and 'id' in v:
                place_qids.add(v['id'])

place_labels = {}
plist = list(place_qids)
total_p = (len(plist) + BATCH_SIZE - 1) // BATCH_SIZE
for i in range(0, len(plist), BATCH_SIZE):
    batch = plist[i:i+BATCH_SIZE]
    bn = i // BATCH_SIZE + 1
    print(f'  Fetching label batch {bn}/{total_p} ...')
    ents = fetch_entities_batch(batch)
    for qid, e in ents.items():
        lbl = get_entity_label(e)
        if lbl:
            place_labels[qid] = lbl
    time.sleep(BATCH_DELAY)

# ---------------------------------------------------------------------------
# Step 4 -- filter: exclude confirmed non-humans; keep unknowns
# ---------------------------------------------------------------------------

excluded = []
for nome in list(qid_map.keys()):
    qid = qid_map[nome]
    ent = entities.get(qid, {})
    human = is_human(ent.get('claims', {}))
    if human is False:   # confirmed non-human
        excluded.append((nome, qid))
        del qid_map[nome]
        # don't add to unfound; these are places/orgs that got wrong search result

if excluded:
    print(f'Confirmed non-human excluded: {len(excluded)}')
    for n, q in excluded[:10]:
        print(f'  {n} -> {q}')
    if len(excluded) > 10:
        print(f'  ... and {len(excluded)-10} more')

# ---------------------------------------------------------------------------
# Step 5 -- append rows to XLSX
# ---------------------------------------------------------------------------

# Find last used row in UOMINI
last_m = 0
for r in ws_m.iter_rows(values_only=True):
    if r[0]:
        last_m += 1

appended_m, appended_f = 0, 0
nd_count = 0

for nome, gender in all_missing:
    qid = qid_map.get(nome)
    search_name = to_search_name(nome)

    if qid and qid in entities and entities[qid].get('claims'):
        ent = entities[qid]
        wikidata_name = get_entity_label(ent) or search_name
        occupation, birth, death = extract_person_data(ent, place_labels)
    else:
        wikidata_name = search_name
        occupation, birth, death = 'N.D.', 'N.D.', 'N.D.'
        nd_count += 1

    if gender == 'M':
        last_m += 1
        ws_m.cell(last_m, 1, wikidata_name)
        ws_m.cell(last_m, 2, 'M')
        ws_m.cell(last_m, 3, occupation)
        ws_m.cell(last_m, 4, birth)
        ws_m.cell(last_m, 5, death)
        appended_m += 1
    else:
        ws_f.append([wikidata_name, 'F', occupation, birth, death])
        appended_f += 1

wb.save(XLSX_IN)
print(f'\nXLSX saved.')
print(f'  Appended -> Uomini: {appended_m}, Donne: {appended_f}')
print(f'  With N.D. (not found or fetch failed): {nd_count}')

# Final counts
ws_m2 = openpyxl.load_workbook(XLSX_IN)['UOMINI']
ws_f2 = openpyxl.load_workbook(XLSX_IN)['DONNE']
total_m = sum(1 for r in ws_m2.iter_rows(values_only=True) if r[0])
total_f = sum(1 for r in ws_f2.iter_rows(min_row=2, values_only=True) if r[0])
print(f'  Total XLSX -> Uomini: {total_m}, Donne: {total_f}')

# ---------------------------------------------------------------------------
# Step 6 -- write not-found report
# ---------------------------------------------------------------------------

with open(NOTFOUND, 'w', encoding='utf-8') as f:
    f.write('Persone non trovate su Wikidata -- da compilare manualmente\n')
    f.write('=' * 60 + '\n\n')
    for nome, gender in unfound:
        f.write(f'[{gender}] {nome}\n')
    if excluded:
        f.write('\nRisultati non umani (probabile match errato):\n')
        for n, q in excluded:
            f.write(f'[?] {n} -> {q}\n')

print(f'  Not-found list -> {NOTFOUND.name}')
print('\nDone.')
