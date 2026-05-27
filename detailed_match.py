import openpyxl, csv, unicodedata, re

PREFIXES = [
    'PASSAGGIO ', 'PIAZZALE ', 'PIAZZETTA ', 'PIAZZA ', 'VIALE ', 'VIA ',
    'ROTONDA ', 'GALLERIA ', 'PONTE ', 'LARGO ', "LOCALITA' ", 'MURA ',
    'SOTTOPASSO ', 'VICOLO ', 'SALITA '
]

def strip_prefix(s):
    for p in PREFIXES:
        if s.startswith(p):
            return s[len(p):]
    return s

def normalize(s):
    if s is None:
        return ''
    s = str(s).strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = s.replace("'", '').replace('`', '').replace('.', '').replace('-', ' ')
    s = strip_prefix(s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

wb = openpyxl.load_workbook('F_M DATE.xlsx')

# Load XLSX uomini
ws_u = wb['UOMINI']
xlsx_m = {}
for row in ws_u.iter_rows(min_row=1, values_only=True):
    if row[0]:
        xlsx_m[normalize(str(row[0]))] = str(row[0])

# Load XLSX donne
ws_d = wb['DONNE']
xlsx_f = {}
for row in ws_d.iter_rows(min_row=2, values_only=True):
    if row[0] and str(row[0]) != 'Nome Personaggio':
        xlsx_f[normalize(str(row[0]))] = str(row[0])

# Load CSV
csv_male = []
csv_female = []
with open('bologna_KG_ready.csv', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        g = row['GENERE']
        nome = row['NOME_PULITO'].strip()
        if g == 'Male':
            csv_male.append(nome)
        elif g == 'Female':
            csv_female.append(nome)

print('=' * 60)
print('RIEPILOGO CONTEGGI')
print('=' * 60)
print(f'KG CSV  — Male: {len(csv_male)},  Female: {len(csv_female)},  Totale: {len(csv_male)+len(csv_female)}')
print(f'XLSX    — Uomini: {len(xlsx_m)},  Donne: {len(xlsx_f)},  Totale: {len(xlsx_m)+len(xlsx_f)}')
print(f'Differenza — M: {len(csv_male)-len(xlsx_m)},  F: {len(csv_female)-len(xlsx_f)}')
print()

# Match function: exact or near match
def find_match(norm, lookup):
    if norm in lookup:
        return lookup[norm]
    for k in lookup:
        if norm == k:
            return lookup[k]
        # Allow one word difference for long names
        if len(norm) > 8 and (norm in k or k in norm):
            return lookup[k]
    return None

# DONNE check
print('=' * 60)
print('DONNE (Female)')
print('=' * 60)
f_found = 0
f_missing = []
for nome in sorted(csv_female):
    norm = normalize(nome)
    match = find_match(norm, xlsx_f)
    if match:
        f_found += 1
    else:
        f_missing.append(nome)

print(f'  Trovate: {f_found}/{len(csv_female)}')
print(f'  Mancanti nel XLSX ({len(f_missing)}):')
for n in f_missing:
    print(f'    - {n}')

# Check XLSX donne not in CSV
xlsx_f_not_in_csv = []
for k, v in xlsx_f.items():
    found = find_match(k, {normalize(n): n for n in csv_female})
    if not found:
        xlsx_f_not_in_csv.append(v)
if xlsx_f_not_in_csv:
    print(f'  Nel XLSX ma NON nel CSV: {xlsx_f_not_in_csv}')
print()

# UOMINI check
print('=' * 60)
print('UOMINI (Male)')
print('=' * 60)
m_found = 0
m_missing_individual = []
m_missing_collective = []
collective_keywords = ['FRATELLI', 'RAGAZZI DEL', 'DECORATI', 'BRIGATE', 'BRIGATA']

for nome in sorted(csv_male):
    norm = normalize(nome)
    match = find_match(norm, xlsx_m)
    if match:
        m_found += 1
    else:
        is_collective = any(kw in nome for kw in collective_keywords)
        if is_collective:
            m_missing_collective.append(nome)
        else:
            m_missing_individual.append(nome)

print(f'  Trovati: {m_found}/{len(csv_male)}')
print(f'  Collettivi senza voce (normale): {len(m_missing_collective)}')
print(f'  Individui mancanti: {len(m_missing_individual)}')
print()
print(f'  Lista completa individui mancanti ({len(m_missing_individual)}):')
for n in m_missing_individual:
    print(f'    - {n}')

# Check XLSX uomini not in CSV
print()
csv_male_norms = {normalize(n): n for n in csv_male}
xlsx_m_not_in_csv = []
for k, v in xlsx_m.items():
    found = find_match(k, csv_male_norms)
    if not found:
        xlsx_m_not_in_csv.append(v)
if xlsx_m_not_in_csv:
    print(f'  Nel XLSX ma NON nel CSV ({len(xlsx_m_not_in_csv)}):')
    for n in xlsx_m_not_in_csv:
        print(f'    - {n}')
